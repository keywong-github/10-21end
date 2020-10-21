import openpyxl

class ReadExcel1:
    def read_excel(self):
        excel=openpyxl.load_workbook('F:\\vscode\myvscode\my-unittest-up\data.xlsx')
        sheet=excel['Sheet1']
        maxrow=sheet.max_row
        maxcolumn=sheet.max_column
        listdata=[]

        for x in range(2,maxrow+1):
            dict1={}
            for y in range(1,maxcolumn+1):
                dict1[sheet.cell(1,y).value]=sheet.cell(x,y).value
            listdata.append(dict1)
        #print(listdata)

        return listdata

ReadExcel1().read_excel()
